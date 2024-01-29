import openpyxl
import pprint


def read_excel_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []

    # Skip rows until the first row with data headers
    row_index = 5
    while not sheet.cell(row=row_index, column=1).value:
        row_index += 1

    headers = [cell.value for cell in sheet[row_index]]

    for row in sheet.iter_rows(min_row=row_index + 3, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)

    workbook.close()
    return data[0]


# Read Excel data
excel_data = read_excel_data("GPATemplate.xlsx", "GPA")

beneficiaries_dict = {}
family_dict = {}

for i in range(1, 5):
    key_fist_name_original = f'ben{i}FirstName*'
    key_last_name_original = f'ben{i}LastName*'
    key_middle_name_original = f'ben{i}MiddleName'
    key_relationship_original = f'ben{i}Relationship*'

    key_first_name_uniformed = f'ben{i}FistName'
    key_last_name_uniformed = f'ben{i}LastName'
    key_middle_name_uniformed = f'ben{i}MiddleName'
    key_relationship_uniformed = f'ben{i}Relationship'

    beneficiaries_dict[key_first_name_uniformed] = excel_data.get(key_fist_name_original, None)
    beneficiaries_dict[key_last_name_uniformed] = excel_data.get(key_last_name_original, None)
    beneficiaries_dict[key_middle_name_uniformed] = excel_data.get(key_middle_name_original, None)
    beneficiaries_dict[key_relationship_uniformed] = excel_data.get(key_relationship_original, None)

for i in range(1, 5):
    key_cs_first_name = f'cs{i}FirstName'
    key_cs_last_name = f'cs{i}LastName'
    key_cs_middle_name = f'cs{i}MiddleName'
    key_cs_birthdate = f'cs{i}BirthDate'

    family_dict[key_cs_first_name] = excel_data.get(key_cs_first_name, None)
    family_dict[key_cs_last_name] = excel_data.get(key_cs_last_name, None)
    family_dict[key_cs_middle_name] = excel_data.get(key_cs_middle_name, None)
    family_dict[key_cs_birthdate] = excel_data.get(key_cs_birthdate, None)

for i in range(1, 3):
    key_sp_first_name = f'sp{i}FirstName'
    key_sp_last_name = f'sp{i}LastName'
    key_sp_middle_name = f'sp{i}MiddleName'
    key_sp_birthdate = f'sp{i}BirthDate'

    family_dict[key_sp_first_name] = excel_data.get(key_sp_first_name, None)
    family_dict[key_sp_last_name] = excel_data.get(key_sp_last_name, None)
    family_dict[key_sp_middle_name] = excel_data.get(key_sp_middle_name, None)
    family_dict[key_sp_birthdate] = excel_data.get(key_sp_birthdate, None)

excel_data['beneficiaries'] = beneficiaries_dict
excel_data['family'] = family_dict

for i in range(1, 5):
    key_first_name_with_star = f'ben{i}FirstName*'
    key_first_name_without_star = f'ben{i}FirstName'
    key_last_name_with_star = f'ben{i}LastName*'
    key_last_name_without_star = f'ben{i}LastName'
    key_middle_name = f'ben{i}MiddleName'
    key_relationship_with_star = f'ben{i}Relationship*'
    key_relationship_without_star = f'ben{i}Relationship'

    key_cs_first_name = f'cs{i}FirstName'
    key_cs_last_name = f'cs{i}LastName'
    key_cs_middle_name = f'cs{i}MiddleName'
    key_cs_birthdate = f'cs{i}BirthDate'

    key_sp_first_name = f'sp{i}FirstName'
    key_sp_last_name = f'sp{i}LastName'
    key_sp_middle_name = f'sp{i}MiddleName'
    key_sp_birthdate = f'sp{i}BirthDate'

    # Check if the key exists before deleting
    if key_first_name_with_star in excel_data:
        del excel_data[key_first_name_with_star]
    if key_first_name_without_star in excel_data:
        del excel_data[key_first_name_without_star]
    if key_last_name_with_star in excel_data:
        del excel_data[key_last_name_with_star]
    if key_last_name_without_star in excel_data:
        del excel_data[key_last_name_without_star]
    if key_middle_name in excel_data:
        del excel_data[key_middle_name]
    if key_relationship_with_star in excel_data:
        del excel_data[key_relationship_with_star]
    if key_relationship_without_star in excel_data:
        del excel_data[key_relationship_without_star]

    del excel_data[key_cs_first_name]
    del excel_data[key_cs_last_name]
    del excel_data[key_cs_middle_name]
    del excel_data[key_cs_birthdate]

    if key_sp_first_name in excel_data:
        del excel_data[key_sp_first_name]
    if key_sp_last_name in excel_data:
        del excel_data[key_sp_last_name]
    if key_sp_middle_name in excel_data:
        del excel_data[key_sp_middle_name]
    if key_sp_birthdate in excel_data:
        del excel_data[key_sp_birthdate]

excel_data['province'] = f'''{excel_data['cityName*']}, {excel_data['provinceName*']}'''

# Print the updated data
pprint.pprint(excel_data)
